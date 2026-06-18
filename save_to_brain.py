import os
import openpyxl
import sqlite3
import re
from datetime import datetime

def save_excel_to_database():
    print("==================================================")
    print(" 🧠 大腦寫入引擎：將 Excel 數據同步至資料庫 v1.1 ")
    print("==================================================")
    
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
    EXCEL_FILE = os.path.join(desktop_path, "Book.xlsx")
    DB_FILE = os.path.join(desktop_path, "stock_brain.db")
    
    if not os.path.exists(EXCEL_FILE):
        print(f"❌ 錯誤：找不到桌面的 Book.xlsx！")
        return
    if not os.path.exists(DB_FILE):
        print(f"❌ 錯誤：找不到大腦資料庫，請先執行 app_backend.py！")
        return
        
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb.worksheets[0]
    
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    today_str = datetime.now().strftime("%Y-%m-%d")
    row_index = 2
    insert_count = 0
    
    print(f"📊 正在掃描 Excel 並將數據寫入資料庫歷史紀錄 ({today_str})...")
    
    while True:
        cell_a = ws[f"A{row_index}"].value
        cell_b = ws[f"B{row_index}"].value
        
        if cell_a is None and cell_b is None:
            if ws[f"A{row_index+1}"].value is None and ws[f"A{row_index+2}"].value is None:
                break
                
        if cell_a is not None:
            stock_id = str(cell_a).strip().split('.')[0]
            stock_name = str(cell_b).strip() if cell_b else "未知"
            
            val_f = ws[f"F{row_index}"].value  # 目前現價
            val_q = ws[f"Q{row_index}"].value  # 5MA
            val_m = ws[f"M{row_index}"].value  # 季線
            
            try:
                # 🎯【核心修正】將撈出來的數字精準保留一位小數
                price = round(float(val_f), 1) if val_f is not None else None
                ma5 = round(float(val_q), 1) if val_q is not None else None
                ma60 = round(float(val_m), 1) if val_m is not None else None
                
                if price is not None:
                    cursor.execute('''
                        INSERT OR REPLACE INTO daily_prices (stock_id, stock_name, trade_date, close_price, ma5, ma60)
                        VALUES (?, ?, ?, ?, ?, ?)
                    ''', (stock_id, stock_name, today_str, price, ma5, ma60))
                    
                    insert_count += 1
            except (ValueError, TypeError):
                pass
                
        row_index += 1
        
    conn.commit()
    conn.close()
    
    print("==================================================")
    print(f"🎉 同步完畢！成功將 {insert_count} 檔一小數股票數據更新進大腦資料庫！")
    print("==================================================")

if __name__ == "__main__":
    save_excel_to_database()